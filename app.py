# app.py

from flask import Flask, render_template, redirect, url_for, request, Response
from modelos import analizar_convenio_ia, cargar_convenios_desde_db, cargar_evaluaciones_desde_google_sheets, EstadoConvenio, Convenio, get_db_connection
from datetime import date, datetime
from fpdf import FPDF
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import io
import unicodedata
# import csv # Ya no es necesario

# Inicializa la aplicación Flask
app = Flask(__name__)

def normalize_text(text: str) -> str:
    """
    Elimina tildes y convierte el texto a minúsculas para una búsqueda insensible.
    """
    if not text:
        return ""
    nfkd_form = unicodedata.normalize('NFD', text)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)]).lower()

class PDF(FPDF):
    def header(self):
        # CORRECCIÓN: Usar fuentes base como 'Helvetica' para evitar warnings.
        self.set_font('Helvetica', 'B', 15)
        self.cell(0, 10, 'Reporte de Estado de Convenios', align='C')
        self.ln(5)
        self.set_font('Helvetica', '', 10)
        self.cell(0, 10, f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}', align='C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', align='C')

@app.route('/') # Define la ruta principal de la web
def dashboard_convenios():
    """
    Carga los convenios, los analiza y los muestra en una plantilla HTML.
    """
    # 1. Obtener los parámetros de filtro desde la URL
    search_query = request.args.get('q', '')
    status_filter = request.args.get('status', '')
    sort_by = request.args.get('sort_by', 'criticidad') # Ordenar por criticidad por defecto
    order = request.args.get('order', 'desc') # Descendente por defecto

    # 2. Cargar todos los convenios para los cálculos globales
    convenios = cargar_convenios_desde_db()

    # Cargar las evaluaciones
    # ¡IMPORTANTE! Reemplaza "Nombre de tu Hoja de Cálculo" con el nombre exacto de tu archivo en Google Drive.
    evaluaciones = cargar_evaluaciones_desde_google_sheets("Calificación de práctica por convenio")
    
    # 3. Calcular KPIs y Notificaciones (sobre la lista completa)
    kpis = {
        'total': len(convenios),
        'vigentes': 0,
        'por_vencer': 0,
        'vencidos': 0
    }
    notificaciones = []
    convenios_analizados = [] # Analizamos todos los convenios una sola vez
    for conv in convenios:
        analisis = analizar_convenio_ia(conv)
        convenios_analizados.append((conv, analisis))

        if analisis.estado == EstadoConvenio.ROJO and analisis.dias_restantes < 0:
            kpis['vencidos'] += 1
        elif analisis.estado == EstadoConvenio.ROJO or analisis.estado == EstadoConvenio.AMARILLO:
            kpis['por_vencer'] += 1
        else:
            kpis['vigentes'] += 1
        
        # Lógica de notificación: convenios que vencen en los próximos 30 días
        if 0 <= analisis.dias_restantes <= 30:
            notificaciones.append(f"¡Atención! El convenio '{conv.nombre}' vence en {analisis.dias_restantes} días.")

    # 4. Aplicar filtros
    convenios_para_tabla = convenios_analizados
    if search_query:
        normalized_query = normalize_text(search_query)
        convenios_para_tabla = [
            (conv, analisis) for conv, analisis in convenios_para_tabla 
            if normalized_query in normalize_text(conv.nombre) or normalized_query in normalize_text(conv.entidad)
        ]
    if status_filter:
        convenios_para_tabla = [(conv, analisis) for conv, analisis in convenios_para_tabla if analisis.estado.value == status_filter]
    
    # 5. Aplicar ordenamiento
    reverse_order = (order == 'desc')
    if sort_by == 'criticidad':
        convenios_para_tabla.sort(key=lambda item: item[1].criticidad, reverse=reverse_order)
    elif sort_by == 'fecha':
        convenios_para_tabla.sort(key=lambda item: item[0].fecha_vencimiento, reverse=reverse_order)
    elif sort_by == 'estado':
        # Asignamos un peso a cada estado para poder ordenarlos
        estado_orden = {EstadoConvenio.ROJO: 3, EstadoConvenio.AMARILLO: 2, EstadoConvenio.VERDE: 1}
        convenios_para_tabla.sort(key=lambda item: estado_orden.get(item[1].estado, 0), reverse=reverse_order)
    
    # 6. Renderizar la plantilla HTML, pasándole todos los datos, incluidos los de ordenamiento
    return render_template('index.html', convenios=convenios_para_tabla, kpis=kpis, notificaciones=notificaciones,
                           evaluaciones=evaluaciones, search_query=search_query, status_filter=status_filter,
                           sort_by=sort_by, order=order)

@app.route('/eliminar/<int:convenio_id>')
def eliminar_convenio(convenio_id):
    """
    Elimina un convenio y sus evaluaciones asociadas de la base de datos.
    """
    conn = get_db_connection()
    conn.execute('DELETE FROM convenios WHERE id = ?', (convenio_id,))
    conn.execute('DELETE FROM evaluaciones WHERE id_convenio = ?', (convenio_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('dashboard_convenios'))

@app.route('/renovar/<int:convenio_id>')
def renovar_convenio(convenio_id):
    """
    Renueva un convenio (añade 1 año a su fecha de vencimiento) en la base de datos.
    """
    conn = get_db_connection()
    convenio = conn.execute('SELECT fecha_vencimiento FROM convenios WHERE id = ?', (convenio_id,)).fetchone()
    if convenio:
        fecha_actual = datetime.strptime(convenio['fecha_vencimiento'], '%Y-%m-%d').date()
        nueva_fecha = fecha_actual.replace(year=fecha_actual.year + 1)
        conn.execute('UPDATE convenios SET fecha_vencimiento = ? WHERE id = ?', (nueva_fecha.strftime('%Y-%m-%d'), convenio_id))
        conn.commit()
    conn.close()
    return redirect(url_for('dashboard_convenios'))

@app.route('/crear', methods=['GET', 'POST'])
def crear_convenio():
    """
    Muestra un formulario para crear un nuevo convenio y lo guarda en la base de datos.
    """
    if request.method == 'POST':
        # Obtener datos del formulario
        nombre = request.form['nombre']
        entidad = request.form['entidad']
        fecha_vencimiento = request.form['fecha_vencimiento']
        responsable = request.form['responsable']
        tipo = request.form['tipo']

        # Insertar en la base de datos
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO convenios (nombre, entidad, fecha_vencimiento, responsable, tipo, renovaciones)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (nombre, entidad, fecha_vencimiento, responsable, tipo, 0)) # Se asume 0 renovaciones al crear
        conn.commit()
        conn.close()
        
        return redirect(url_for('dashboard_convenios'))

    return render_template('crear_convenio.html')

@app.route('/export/pdf')
def export_pdf():
    """
    Genera un reporte en PDF con la lista de todos los convenios.
    """
    convenios = cargar_convenios_desde_db()
    
    # CORRECCIÓN: Simplificar la inicialización del PDF.
    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page(orientation='L') # 'L' para Landscape (apaisado)
    
    # Definir cabeceras y anchos de columna
    headers = ['Estado', 'Criticidad', 'Nombre', 'Entidad', 'Vencimiento', 'Días Rest.', 'Responsable']
    col_widths = [25, 25, 70, 55, 25, 20, 50] # Ajustamos anchos para A4

    # Escribir cabeceras
    pdf.set_font('Helvetica', 'B', 8)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, align='C')
    pdf.ln()

    # Escribir datos de la tabla
    pdf.set_font('Helvetica', '', 8)
    for convenio in convenios:
        analisis = analizar_convenio_ia(convenio)
        # CORRECCIÓN: Codificar el texto a 'latin-1' con 'replace' ANTES de pasarlo a la celda.
        # Esto previene la corrupción del archivo PDF por caracteres especiales.
        pdf.cell(col_widths[0], 10, analisis.estado.value, border=1, align='C')
        pdf.cell(col_widths[1], 10, str(analisis.criticidad), border=1, align='C')
        pdf.cell(col_widths[2], 10, convenio.nombre.encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.cell(col_widths[3], 10, convenio.entidad.encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.cell(col_widths[4], 10, convenio.fecha_vencimiento.strftime('%d-%m-%Y'), border=1, align='C')
        pdf.cell(col_widths[5], 10, str(analisis.dias_restantes), border=1, align='C')
        pdf.cell(col_widths[6], 10, convenio.responsable.encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.ln()

    # CORRECCIÓN FINAL: Asegurarse de que la salida sea explícitamente del tipo 'bytes'
    # que el servidor Werkzeug espera.
    return Response(bytes(pdf.output()),
                    mimetype='application/pdf',
                    headers={'Content-Disposition':'attachment;filename=reporte_convenios.pdf'})

@app.route('/export/excel')
def export_excel():
    """
    Genera un reporte en formato Excel (.xlsx) con la lista de todos los convenios.
    """
    convenios = cargar_convenios_desde_db()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Convenios"

    # Escribir cabeceras
    headers = ['ID', 'Estado', 'Criticidad (IA)', 'Nombre del Convenio', 'Entidad', 'Fecha de Vencimiento', 'Días Restantes', 'Responsable', 'Tipo', 'Renovaciones']
    ws.append(headers)

    # CORRECCIÓN: Ajustar el ancho de las columnas automáticamente
    column_widths = []
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = len(header) + 5

    # Escribir datos de la tabla
    for convenio in convenios:
        analisis = analizar_convenio_ia(convenio)
        ws.append([
            convenio.id, analisis.estado.value, analisis.criticidad, convenio.nombre, convenio.entidad,
            convenio.fecha_vencimiento, analisis.dias_restantes, convenio.responsable, convenio.tipo, convenio.renovaciones
        ])

    # Guardar el archivo en un buffer de memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition':'attachment;filename=reporte_convenios.xlsx'})

if __name__ == '__main__':
    app.run(debug=True) # Inicia el servidor en modo de depuración